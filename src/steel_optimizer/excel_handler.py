from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import (
    Coil,
    InputData,
    Order,
    Settings,
)


def _to_bool(value: Any) -> bool:
    """
    Convert common Excel values into a boolean.
    """

    if isinstance(value, bool):

        return value

    if isinstance(value, str):

        return (
            value.strip().lower()
            in {
                "yes",
                "true",
                "1",
                "y",
            }
        )

    if isinstance(
        value,
        (int, float),
    ):

        return value != 0

    return False


def read_input_excel(
    file_path: str | Path,
) -> InputData:
    """
    Read all optimizer inputs from Excel.

    Expected sheets:

    Coil:
        A: Parameter
        B: Value

    Orders:
        A: Width (mm)
        B: Required Weight (kg)

    StockSizes:
        A: Width (mm)

    Settings:
        A: Setting
        B: Value
    """

    file_path = Path(
        file_path
    )

    if not file_path.exists():

        raise FileNotFoundError(
            f"Input Excel file not found: "
            f"{file_path}"
        )

    workbook = load_workbook(
        filename=file_path,
        data_only=True,
    )

    coil = _read_coil(
        workbook
    )

    orders = _read_orders(
        workbook
    )

    stock_widths = (
        _read_stock_widths(
            workbook
        )
    )

    settings = _read_settings(
        workbook
    )

    return InputData(
        coil=coil,
        orders=orders,
        stock_widths_mm=stock_widths,
        settings=settings,
    )


def _read_coil(
    workbook,
) -> Coil:

    if "Coil" not in workbook.sheetnames:

        raise ValueError(
            "Missing required sheet: Coil"
        )

    sheet = workbook["Coil"]

    values = {}

    for parameter, value in sheet.iter_rows(
        min_row=2,
        values_only=True,
    ):

        if parameter is None:

            continue

        key = str(
            parameter
        ).strip().lower()

        values[key] = value

    weight_value = values.get(
        "weight (kg)"
    )

    if weight_value is None:

        weight_kg = None

    else:

        weight_kg = float(
            weight_value
        )

    return Coil(

        thickness_mm=float(
            values["thickness (mm)"]
        ),

        width_mm=int(
            values["width (mm)"]
        ),

        weight_kg=weight_kg,

        kerf_mm=float(
            values["kerf (mm)"]
        ),
    )


def _read_orders(
    workbook,
) -> list[Order]:

    if "Orders" not in workbook.sheetnames:

        raise ValueError(
            "Missing required sheet: Orders"
        )

    sheet = workbook["Orders"]

    orders = []

    for width, required_weight in (
        sheet.iter_rows(
            min_row=2,
            values_only=True,
        )
    ):

        if width is None:

            continue

        if required_weight is None:

            continue

        orders.append(

            Order(

                width_mm=int(
                    width
                ),

                required_weight_kg=float(
                    required_weight
                ),
            )
        )

    return orders


def _read_stock_widths(
    workbook,
) -> list[int]:

    if (
        "StockSizes"
        not in workbook.sheetnames
    ):

        raise ValueError(
            "Missing required sheet: "
            "StockSizes"
        )

    sheet = workbook[
        "StockSizes"
    ]

    stock_widths = []

    for (
        width,
    ) in sheet.iter_rows(
        min_row=2,
        values_only=True,
    ):

        if width is None:

            continue

        stock_widths.append(
            int(width)
        )

    return sorted(
        set(
            stock_widths
        )
    )


def _read_settings(
    workbook,
) -> Settings:

    if (
        "Settings"
        not in workbook.sheetnames
    ):

        raise ValueError(
            "Missing required sheet: "
            "Settings"
        )

    sheet = workbook[
        "Settings"
    ]

    values = {}

    for setting, value in (
        sheet.iter_rows(
            min_row=2,
            values_only=True,
        )
    ):

        if setting is None:

            continue

        key = str(
            setting
        ).strip().lower()

        values[key] = value

    return Settings(

        allow_overproduction=_to_bool(
            values[
                "allow overproduction"
            ]
        ),

        max_overproduction_percent=float(
            values[
                "max overproduction (%)"
            ]
        ),

        max_knives=int(
            values[
                "max knives"
            ]
        ),

        allow_stock_production=_to_bool(
            values[
                "allow stock production"
            ]
        ),
    )